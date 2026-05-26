"""
Convert each sheet of an Excel workbook into a CSV file.

Used to derive the manually-entered ground-truth CSVs (one per week) from the
original Excel workbooks of the Canada Notifiable Disease Dataset.

Usage:
    python xlsx_to_csvs.py <input_xlsx> <output_dir>

Example (run from repo root):
    python extraction/xlsx_to_csvs.py \\
        data/source_xlsx/cdi_ca_1956_wk_prov_dbs.xlsx data/manual/1956

Requires:
    pip install openpyxl
"""

import csv
import argparse
from pathlib import Path
from openpyxl import load_workbook


def safe_name(name: str) -> str:
    """Strip filesystem-illegal characters from a sheet name."""
    for ch in '\\/:*?"<>|':
        name = name.replace(ch, "_")
    return name.strip()[:120] or "Sheet"


def sheet_to_csv(ws, out_path: Path):
    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            # keep empty cells as empty strings so columns don't shift
            w.writerow(["" if v is None else v for v in row])


def xlsx_to_csvs(input_xlsx: str, output_dir: str):
    input_path = Path(input_xlsx)
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(input_path, data_only=True)
    for i, ws in enumerate(wb.worksheets, 1):
        out = out_dir / f"{input_path.stem}__{i:04d}__{safe_name(ws.title)}.csv"
        sheet_to_csv(ws, out)
        print(f"Wrote {out}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Split an Excel workbook into one CSV per sheet."
    )
    parser.add_argument("input_xlsx", help="Path to input .xlsx file.")
    parser.add_argument("output_dir", help="Directory to write CSV files.")
    args = parser.parse_args()
    xlsx_to_csvs(args.input_xlsx, args.output_dir)
