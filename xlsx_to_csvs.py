import csv
from pathlib import Path
from openpyxl import load_workbook

def safe_name(name: str) -> str:
    bad = '\\/:*?"<>|'
    for ch in bad:
        name = name.replace(ch, "_")
    return name.strip()[:120] or "Sheet"

def sheet_to_csv(ws, out_path: Path):
    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            # keep empty cells as empty strings so columns don't shift
            w.writerow(["" if v is None else v for v in row])

def xlsx_to_csvs(input_xlsx: str, output_dir: str):
    input_path = Path(input_xlsx)
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(input_path, data_only=True)  # values (not formulas)
    for i, ws in enumerate(wb.worksheets, 1):
        out = out_dir / f"{input_path.stem}__{i:04d}__{safe_name(ws.title)}.csv"
        sheet_to_csv(ws, out)
        print("Wrote", out)

if __name__ == "__main__":
    xlsx_to_csvs(
        r"1958_Manual\cdi_ca_1958_wk_prov_dbs.xlsx",
        r"1958_Manual_csv"
    )
